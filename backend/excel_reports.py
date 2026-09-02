"""Laporan Excel PT SJM — semua workbook dibangun di sini agar rapi & konsisten.

Prinsip:
- Header hijau korporat, judul + subjudul, freeze panes, border tipis, zebra baris.
- Angka pakai number_format Indonesia (Rp / Kg) dan TOTAL memakai RUMUS (=SUM) agar tetap editable.
- Nama sheet aman (max 31 char, tanpa karakter ilegal Excel).
"""
from __future__ import annotations

import io
import re
from datetime import date as _date

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

GREEN = '15803D'
GREEN_DARK = '0F5132'
GREEN_SOFT = 'E9F5ED'
GOLD = 'B97819'
GREY = '6B7280'
RED = 'B42318'

FMT_RP = 'Rp #,##0'
FMT_KG = '#,##0'
FMT_TON = '#,##0.00'
FMT_PCT = '0.00"%"'

THIN = Side(style='thin', color='D3DCD5')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HARI = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
BULAN = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 'Juli',
         'Agustus', 'September', 'Oktober', 'November', 'Desember']


# ---------------------------------------------------------------- helpers
def d_id(s: str) -> str:
    """2026-09-01 -> 01/09/2026"""
    try:
        y, m, d = s.split('-')
        return f'{d}/{m}/{y}'
    except Exception:
        return s or '-'


def hari_id(s: str) -> str:
    try:
        y, m, d = (int(x) for x in s.split('-'))
        return HARI[_date(y, m, d).weekday()]
    except Exception:
        return '-'


def bulan_id(month: str) -> str:
    try:
        y, m = month.split('-')
        return f'{BULAN[int(m) - 1]} {y}'
    except Exception:
        return month or '-'


def periode_label(start: str | None, end: str | None, dates: list[str]) -> str:
    if start and end:
        return f'{d_id(start)} — {d_id(end)}'
    if dates:
        lo, hi = min(dates), max(dates)
        return f'{d_id(lo)} — {d_id(hi)}' if lo != hi else d_id(lo)
    return 'Semua transaksi batch'


def safe_sheet(name: str, used: set[str]) -> str:
    base = re.sub(r'[\[\]\*/\\\?:]', ' ', (name or 'Sheet')).strip()[:31] or 'Sheet'
    out, i = base, 2
    while out.lower() in used:
        suffix = f' ({i})'
        out = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(out.lower())
    return out


def title_block(ws, title: str, subtitle: str, ncols: int, extra: str = '') -> int:
    """Judul laporan. Return nomor baris untuk header tabel."""
    last = get_column_letter(max(ncols, 3))
    ws.merge_cells(f'A1:{last}1')
    c = ws['A1']
    c.value = title
    c.font = Font(bold=True, size=15, color=GREEN_DARK)
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells(f'A2:{last}2')
    c = ws['A2']
    c.value = subtitle
    c.font = Font(size=10, color=GREY)
    row = 3
    if extra:
        ws.merge_cells(f'A3:{last}3')
        c = ws['A3']
        c.value = extra
        c.font = Font(size=10, color=GREY)
        row = 4
    ws.row_dimensions[row].height = 6
    return row + 1


def table_header(ws, row: int, heads: list[str], widths: list[int]) -> None:
    for i, h in enumerate(heads, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=10)
        c.fill = PatternFill('solid', fgColor=GREEN)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[row].height = 30
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def style_rows(ws, first: int, last: int, ncols: int, formats: dict[int, str], zebra=True) -> None:
    for r in range(first, last + 1):
        for i in range(1, ncols + 1):
            c = ws.cell(row=r, column=i)
            c.border = BORDER
            c.font = Font(size=10)
            if i in formats:
                c.number_format = formats[i]
                c.alignment = Alignment(horizontal='right')
            else:
                c.alignment = Alignment(horizontal='left')
            if zebra and (r - first) % 2 == 1:
                c.fill = PatternFill('solid', fgColor='F7FAF8')


def total_row(ws, row: int, values: list, ncols: int, formats: dict[int, str]) -> None:
    for i in range(1, ncols + 1):
        v = values[i - 1] if i - 1 < len(values) else None
        c = ws.cell(row=row, column=i, value=v)
        c.font = Font(bold=True, size=10, color=GREEN_DARK)
        c.fill = PatternFill('solid', fgColor=GREEN_SOFT)
        c.border = BORDER
        if i in formats:
            c.number_format = formats[i]
            c.alignment = Alignment(horizontal='right')
    ws.row_dimensions[row].height = 20


def kv_block(ws, row: int, pairs: list[tuple[str, object, str]], col: int = 1) -> int:
    """Blok ringkasan label/nilai vertikal."""
    for i, (label, value, fmt) in enumerate(pairs):
        lc = ws.cell(row=row + i, column=col, value=label)
        lc.font = Font(size=10, bold=True, color=GREY)
        vc = ws.cell(row=row + i, column=col + 1, value=value)
        vc.font = Font(size=11, bold=True, color=GREEN_DARK)
        if fmt:
            vc.number_format = fmt
        vc.alignment = Alignment(horizontal='left')
    return row + len(pairs)


def as_bytes(wb: Workbook) -> io.BytesIO:
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# ---------------------------------------------------------------- sheets
def sheet_transaksi(ws, txs: list[dict], periode: str) -> None:
    heads = ['Kode', 'Tanggal', 'Hari', 'Jam Terima', 'Jam Keluar', 'Pemilik', 'Kendaraan',
             'Gross (Kg)', 'Tare (Kg)', 'Netto (Kg)', 'Netto (Ton)', 'Harga/Kg', 'Total Beli']
    widths = [17, 12, 10, 11, 11, 18, 12, 12, 11, 12, 11, 12, 16]
    fmts = {8: FMT_KG, 9: FMT_KG, 10: FMT_KG, 11: FMT_TON, 12: FMT_RP, 13: FMT_RP}
    hr = title_block(ws, 'BUKU PENERIMAAN BUAH SAWIT — PT SJM',
                     f'Periode {periode} · {len(txs)} transaksi', len(heads),
                     'Kolom Netto, Ton, dan Total memakai rumus Excel sehingga tetap bisa dikoreksi.')
    table_header(ws, hr, heads, widths)
    txs = sorted(txs, key=lambda t: (t.get('date', ''), t.get('arrival_time', '')))
    r = hr + 1
    for t in txs:
        ws.append([]) if False else None
        row = [t.get('transaction_code'), d_id(t.get('date')), hari_id(t.get('date')),
               t.get('arrival_time'), t.get('exit_time'), t.get('owner_name'), t.get('vehicle_type'),
               t.get('gross_kg', 0), t.get('tare_kg') or 0,
               f'=H{r}-I{r}', f'=ROUND(J{r}/1000,2)', t.get('price_per_kg', 0), f'=J{r}*L{r}']
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if txs:
        total_row(ws, r, ['TOTAL', '', '', '', '', '', '',
                          f'=SUM(H{hr+1}:H{last})', f'=SUM(I{hr+1}:I{last})',
                          f'=SUM(J{hr+1}:J{last})', f'=SUM(K{hr+1}:K{last})', '',
                          f'=SUM(M{hr+1}:M{last})'], len(heads), fmts)
    else:
        ws.cell(row=r, column=1, value='Belum ada transaksi pada periode ini.').font = Font(italic=True, color=GREY)


def _owner_rows(txs: list[dict]) -> dict:
    per = {}
    for t in txs:
        o = per.setdefault(t.get('owner_name', '-'), {
            'name': t.get('owner_name', '-'), 'vehicle': t.get('vehicle_type', '-'),
            'count': 0, 'netto': 0, 'total': 0, 'txs': []})
        o['count'] += 1
        o['netto'] += t.get('netto_kg', 0)
        o['total'] += t.get('total_amount', 0)
        o['txs'].append(t)
    return per


def sheet_rekap_pemilik(ws, txs: list[dict], periode: str) -> None:
    per = _owner_rows(txs)
    heads = ['Pemilik', 'Kendaraan', 'Jumlah Muatan', 'Netto (Kg)', 'Netto (Ton)',
             'Harga Rata-rata/Kg', 'Total Pembelian', 'Kontribusi']
    widths = [22, 14, 15, 13, 12, 18, 18, 13]
    fmts = {3: FMT_KG, 4: FMT_KG, 5: FMT_TON, 6: FMT_RP, 7: FMT_RP, 8: FMT_PCT}
    hr = title_block(ws, 'REKAP PER PEMILIK — PT SJM',
                     f'Periode {periode} · {len(per)} pemilik', len(heads),
                     'Detail tiap pemilik ada pada sheet terpisah sesuai nama pemilik.')
    table_header(ws, hr, heads, widths)
    r = hr + 1
    grand = sum(o['netto'] for o in per.values()) or 1
    for o in sorted(per.values(), key=lambda x: -x['netto']):
        vals = [o['name'], o['vehicle'], o['count'], o['netto'], round(o['netto'] / 1000, 2),
                round(o['total'] / o['netto']) if o['netto'] else 0, o['total'],
                round(o['netto'] / grand * 100, 2)]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if per:
        total_row(ws, r, ['TOTAL', '', f'=SUM(C{hr+1}:C{last})', f'=SUM(D{hr+1}:D{last})',
                          f'=SUM(E{hr+1}:E{last})', '', f'=SUM(G{hr+1}:G{last})', 100],
                  len(heads), fmts)
    else:
        ws.cell(row=r, column=1, value='Belum ada transaksi pada periode ini.').font = Font(italic=True, color=GREY)


def sheet_pemilik_detail(ws, owner: dict, periode: str) -> None:
    heads = ['Tanggal', 'Hari', 'Kode', 'Jam Terima', 'Jam Keluar', 'Gross (Kg)', 'Tare (Kg)',
             'Netto (Kg)', 'Netto (Ton)', 'Harga/Kg', 'Total Harga']
    widths = [12, 10, 17, 11, 11, 12, 11, 12, 11, 12, 17]
    fmts = {6: FMT_KG, 7: FMT_KG, 8: FMT_KG, 9: FMT_TON, 10: FMT_RP, 11: FMT_RP}
    hr = title_block(ws, f"REKAP PEMILIK — {owner['name'].upper()}",
                     f"{owner['vehicle']} · Periode {periode}", len(heads),
                     f"{owner['count']} muatan · Netto {owner['netto']:,} Kg · Nilai Rp {owner['total']:,}".replace(',', '.'))
    table_header(ws, hr, heads, widths)
    r = hr + 1
    for t in sorted(owner['txs'], key=lambda x: (x.get('date', ''), x.get('arrival_time', ''))):
        vals = [d_id(t.get('date')), hari_id(t.get('date')), t.get('transaction_code'),
                t.get('arrival_time'), t.get('exit_time'), t.get('gross_kg', 0), t.get('tare_kg') or 0,
                f'=F{r}-G{r}', f'=ROUND(H{r}/1000,2)', t.get('price_per_kg', 0), f'=H{r}*J{r}']
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if owner['txs']:
        total_row(ws, r, ['TOTAL', '', f'{owner["count"]} muatan', '', '',
                          f'=SUM(F{hr+1}:F{last})', f'=SUM(G{hr+1}:G{last})',
                          f'=SUM(H{hr+1}:H{last})', f'=SUM(I{hr+1}:I{last})', '',
                          f'=SUM(K{hr+1}:K{last})'], len(heads), fmts)
        r += 2
        kv_block(ws, r, [
            ('Rata-rata netto / muatan (Kg)', round(owner['netto'] / owner['count']), FMT_KG),
            ('Harga rata-rata / Kg', round(owner['total'] / owner['netto']) if owner['netto'] else 0, FMT_RP),
            ('Total yang dibayarkan', owner['total'], FMT_RP),
        ])
    ws.cell(row=r + 5, column=1, value='Tanda tangan pemilik').font = Font(size=9, color=GREY)
    ws.cell(row=r + 5, column=8, value='Petugas timbang PT SJM').font = Font(size=9, color=GREY)


def sheet_rekap_harian(ws, txs: list[dict], periode: str) -> None:
    days = {}
    for t in txs:
        g = days.setdefault(t.get('date'), {'date': t.get('date'), 'count': 0, 'netto': 0, 'total': 0, 'owners': set()})
        g['count'] += 1
        g['netto'] += t.get('netto_kg', 0)
        g['total'] += t.get('total_amount', 0)
        g['owners'].add(t.get('owner_name'))
    heads = ['Tanggal', 'Hari', 'Jumlah Muatan', 'Pemilik Terlibat', 'Netto (Kg)', 'Netto (Ton)',
             'Harga Rata-rata/Kg', 'Total Pembelian']
    widths = [13, 11, 15, 16, 13, 12, 18, 18]
    fmts = {3: FMT_KG, 4: FMT_KG, 5: FMT_KG, 6: FMT_TON, 7: FMT_RP, 8: FMT_RP}
    hr = title_block(ws, 'REKAP MUATAN HARIAN — PT SJM',
                     f'Periode {periode} · {len(days)} hari operasi', len(heads),
                     'Ringkasan muatan yang diterima setiap hari beserta nilai pembeliannya.')
    table_header(ws, hr, heads, widths)
    r = hr + 1
    for d in sorted(days.values(), key=lambda x: x['date'] or ''):
        vals = [d_id(d['date']), hari_id(d['date']), d['count'], len(d['owners']), d['netto'],
                round(d['netto'] / 1000, 2),
                round(d['total'] / d['netto']) if d['netto'] else 0, d['total']]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if days:
        total_row(ws, r, ['TOTAL', f'{len(days)} hari', f'=SUM(C{hr+1}:C{last})', '',
                          f'=SUM(E{hr+1}:E{last})', f'=SUM(F{hr+1}:F{last})', '',
                          f'=SUM(H{hr+1}:H{last})'], len(heads), fmts)
        r += 2
        tot_netto = sum(d['netto'] for d in days.values())
        tot_val = sum(d['total'] for d in days.values())
        kv_block(ws, r, [
            ('Rata-rata netto / hari (Kg)', round(tot_netto / len(days)), FMT_KG),
            ('Rata-rata nilai / hari', round(tot_val / len(days)), FMT_RP),
            ('Hari tertinggi', d_id(max(days.values(), key=lambda x: x['netto'])['date']), ''),
        ])
        chart = LineChart()
        chart.title = 'Netto harian (Kg)'
        chart.height, chart.width = 7, 18
        data = Reference(ws, min_col=5, min_row=hr, max_row=last)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f'A{r + 5}')
    else:
        ws.cell(row=r, column=1, value='Belum ada transaksi pada periode ini.').font = Font(italic=True, color=GREY)


def sheet_laba_rugi(ws, fin: dict, periode: str) -> None:
    days = fin.get('days', [])
    tot = fin.get('totals', {})
    heads = ['Tanggal', 'Hari', 'Muatan', 'Netto (Kg)', 'Total Beli', 'Angkut/Kg', 'Total Angkut',
             'Total Modal', 'Tare (Kg)', 'Grading', 'Netto Jual (Kg)', 'Harga SIP/Kg',
             'Harga Jual', 'PPh 22', 'Tarif / Mode PPh', 'Untung / Rugi', 'Status']
    widths = [12, 10, 9, 12, 16, 11, 15, 16, 11, 10, 14, 13, 17, 15, 18, 17, 13]
    fmts = {3: FMT_KG, 4: FMT_KG, 5: FMT_RP, 6: FMT_RP, 7: FMT_RP, 8: FMT_RP, 9: FMT_KG,
            10: FMT_PCT, 11: FMT_KG, 12: FMT_RP, 13: FMT_RP, 14: FMT_RP, 16: FMT_RP}
    hr = title_block(ws, 'LAPORAN LABA RUGI HARIAN — PT SJM',
                     f'Periode {periode} · {tot.get("configured_days", 0)} dari {tot.get("total_days", 0)} hari sudah digrading',
                     len(heads),
                     f'Modal = Total Beli + Angkut · Harga Jual = Netto Jual × Harga PT SIP · Untung = Harga Jual − PPh 22 − Modal · '
                     f'Tarif PPh 22 default {tot.get("pph_rate_default", 0.25)}% (bisa dikelola per tanggal)')
    table_header(ws, hr, heads, widths)
    r = hr + 1
    for d in days:
        conf = d.get('configured')
        tarif = ''
        if conf:
            tarif = (f'Manual · {d.get("pph_rate_pct", 0)}% efektif' if d.get('pph_mode') == 'manual'
                     else f'Otomatis {d.get("pph_rate_pct", 0)}%')
            if d.get('pph_note'):
                tarif += f' · {d["pph_note"]}'
        vals = [d_id(d.get('date')), hari_id(d.get('date')), d.get('tx_count', 0), d.get('netto_kg', 0),
                d.get('total_beli', 0), d.get('freight_per_kg', 0), d.get('total_angkut', 0),
                d.get('total_modal', 0),
                d.get('total_tare_kg') if conf else None, d.get('grading_pct') if conf else None,
                d.get('netto_jual') if conf else None, d.get('sip_price_per_kg') if conf else None,
                d.get('harga_jual') if conf else None, d.get('pph22') if conf else None, tarif or None,
                d.get('untung_rugi') if conf else None, 'Sudah grading' if conf else 'Belum grading']
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value='—' if v is None else v)
        c = ws.cell(row=r, column=16)
        if conf:
            c.font = Font(size=10, bold=True, color=GREEN_DARK if (d.get('untung_rugi') or 0) >= 0 else RED)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if days:
        total_row(ws, r, ['TOTAL', f'{len(days)} hari', f'=SUM(C{hr+1}:C{last})', f'=SUM(D{hr+1}:D{last})',
                          f'=SUM(E{hr+1}:E{last})', '', f'=SUM(G{hr+1}:G{last})', f'=SUM(H{hr+1}:H{last})',
                          tot.get('total_tare_kg', 0), '', tot.get('netto_jual', 0), '',
                          tot.get('harga_jual', 0), tot.get('pph22', 0),
                          f'Efektif {tot.get("pph_effective_pct", 0)}%', tot.get('untung_rugi', 0), ''],
                  len(heads), fmts)
        c = ws.cell(row=r, column=16)
        c.font = Font(bold=True, size=11, color=GREEN_DARK if (tot.get('untung_rugi') or 0) >= 0 else RED)
        r += 2
        margin = (tot.get('untung_rugi', 0) / tot['harga_jual'] * 100) if tot.get('harga_jual') else 0
        head = ws.cell(row=r, column=1, value='RINGKASAN UNTUK MANAJEMEN')
        head.font = Font(bold=True, size=11, color=GREEN_DARK)
        kv_block(ws, r + 1, [
            ('Total netto diterima (Kg)', tot.get('netto_kg', 0), FMT_KG),
            ('Total modal (beli + angkut)', tot.get('total_modal', 0), FMT_RP),
            ('Total harga jual ke PT SIP', tot.get('harga_jual', 0), FMT_RP),
            (f'PPh 22 (efektif {tot.get("pph_effective_pct", 0)}%)', tot.get('pph22', 0), FMT_RP),
            ('Hari dengan PPh dikelola manual', f"{tot.get('pph_manual_days', 0)} hari", ''),
            ('Untung / rugi bersih', tot.get('untung_rugi', 0), FMT_RP),
            ('Margin terhadap harga jual', round(margin, 2), FMT_PCT),
            ('Hari sudah digrading', f"{tot.get('configured_days', 0)} / {tot.get('total_days', 0)}", ''),
        ])
        srow = r + 10
        ws.cell(row=srow, column=1, value='Disiapkan oleh').font = Font(size=9, color=GREY)
        ws.cell(row=srow, column=6, value='Diperiksa oleh').font = Font(size=9, color=GREY)
        ws.cell(row=srow, column=12, value='Disetujui oleh').font = Font(size=9, color=GREY)
        for col in (1, 6, 12):
            ws.cell(row=srow + 4, column=col, value='(………………………………)').font = Font(size=9)
    else:
        ws.cell(row=r, column=1, value='Belum ada transaksi pada periode ini.').font = Font(italic=True, color=GREY)


def sheet_detail_finance(ws, fin: dict, periode: str) -> None:
    txs = fin.get('transactions', [])
    heads = ['Kode', 'Tanggal', 'Pemilik', 'Netto (Kg)', 'Harga/Kg', 'Total Beli', 'Angkut',
             'Tare Grading (Kg)', 'Grading', 'Netto Jual (Kg)', 'Harga Jual']
    widths = [17, 12, 18, 12, 12, 16, 14, 16, 10, 14, 17]
    fmts = {4: FMT_KG, 5: FMT_RP, 6: FMT_RP, 7: FMT_RP, 8: FMT_KG, 9: FMT_PCT, 10: FMT_KG, 11: FMT_RP}
    hr = title_block(ws, 'DETAIL GRADING PER TRANSAKSI — PT SJM',
                     f'Periode {periode} · {len(txs)} transaksi', len(heads),
                     'Tare grading dibagi acak proporsional terhadap netto masing-masing muatan.')
    table_header(ws, hr, heads, widths)
    r = hr + 1
    for t in txs:
        vals = [t.get('code'), d_id(t.get('date')), t.get('owner_name'), t.get('netto_kg', 0),
                t.get('price_per_kg', 0), t.get('total_amount', 0), t.get('freight', 0),
                t.get('tare_kg'), t.get('grading_pct'), t.get('netto_jual'), t.get('harga_jual')]
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value='—' if v is None else v)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts)
    if txs:
        total_row(ws, r, ['TOTAL', '', f'{len(txs)} transaksi', f'=SUM(D{hr+1}:D{last})', '',
                          f'=SUM(F{hr+1}:F{last})', f'=SUM(G{hr+1}:G{last})', '', '', '', ''],
                  len(heads), fmts)
    else:
        ws.cell(row=r, column=1, value='Belum ada transaksi pada periode ini.').font = Font(italic=True, color=GREY)


def sheet_harga(ws, prices: list[dict], month: str | None) -> None:
    heads = ['Tanggal', 'Hari', 'Harga/Kg', 'Naik/Turun (Rp)', 'Naik/Turun (%)', 'Tren', 'Catatan']
    widths = [13, 11, 14, 17, 16, 12, 40]
    fmts = {3: FMT_RP, 4: FMT_RP, 5: FMT_PCT}
    hr = title_block(ws, 'MONITORING HARGA BELI / KG — PT SJM',
                     f'Bulan {bulan_id(month)} · {len(prices)} tanggal tercatat' if month else f'{len(prices)} tanggal tercatat',
                     len(heads), 'Naik/turun dihitung terhadap harga tanggal tercatat sebelumnya.')
    table_header(ws, hr, heads, widths)
    r = hr + 1
    for p in prices:
        ch = p.get('change')
        vals = [d_id(p.get('date')), hari_id(p.get('date')), p.get('price_per_kg', 0),
                ch if ch is not None else '—',
                p.get('change_pct') if p.get('change_pct') is not None else '—',
                ('Naik' if ch > 0 else 'Turun' if ch < 0 else 'Tetap') if ch is not None else 'Awal',
                p.get('note') or '']
        for i, v in enumerate(vals, 1):
            ws.cell(row=r, column=i, value=v)
        if ch:
            ws.cell(row=r, column=4).font = Font(size=10, bold=True, color=GREEN_DARK if ch > 0 else RED)
            ws.cell(row=r, column=6).font = Font(size=10, bold=True, color=GREEN_DARK if ch > 0 else RED)
        r += 1
    last = r - 1
    style_rows(ws, hr + 1, last, len(heads), fmts, zebra=True)
    if prices:
        vals = [p.get('price_per_kg', 0) for p in prices]
        r += 1
        kv_block(ws, r, [
            ('Harga tertinggi', max(vals), FMT_RP),
            ('Harga terendah', min(vals), FMT_RP),
            ('Harga rata-rata', round(sum(vals) / len(vals)), FMT_RP),
            ('Harga terakhir', vals[-1], FMT_RP),
        ])
        chart = LineChart()
        chart.title = 'Tren harga / Kg'
        chart.height, chart.width = 7, 18
        data = Reference(ws, min_col=3, min_row=hr, max_row=last)
        cats = Reference(ws, min_col=1, min_row=hr + 1, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f'A{r + 6}')
    else:
        ws.cell(row=r, column=1, value='Belum ada harga tercatat.').font = Font(italic=True, color=GREY)


# ---------------------------------------------------------------- workbooks
def wb_owners(txs: list[dict], periode: str, owner: str | None = None) -> io.BytesIO:
    wb = Workbook()
    used: set[str] = set()
    per = _owner_rows(txs)
    if owner:
        o = per.get(owner)
        ws = wb.active
        ws.title = safe_sheet(owner, used)
        sheet_pemilik_detail(ws, o or {'name': owner, 'vehicle': '-', 'count': 0, 'netto': 0, 'total': 0, 'txs': []}, periode)
        return as_bytes(wb)
    ws = wb.active
    ws.title = safe_sheet('Ringkasan Pemilik', used)
    sheet_rekap_pemilik(ws, txs, periode)
    for o in sorted(per.values(), key=lambda x: -x['netto']):
        sheet_pemilik_detail(wb.create_sheet(safe_sheet(o['name'], used)), o, periode)
    return as_bytes(wb)


def wb_daily(txs: list[dict], periode: str) -> io.BytesIO:
    wb = Workbook()
    used: set[str] = set()
    sheet_rekap_harian(wb.active, txs, periode)
    wb.active.title = safe_sheet('Rekap Harian', used)
    sheet_transaksi(wb.create_sheet(safe_sheet('Detail Transaksi', used)), txs, periode)
    return as_bytes(wb)


def wb_finance(fin: dict, periode: str) -> io.BytesIO:
    wb = Workbook()
    used: set[str] = set()
    sheet_laba_rugi(wb.active, fin, periode)
    wb.active.title = safe_sheet('Laba Rugi Harian', used)
    sheet_detail_finance(wb.create_sheet(safe_sheet('Detail Grading', used)), fin, periode)
    return as_bytes(wb)


def wb_prices(prices: list[dict], month: str | None) -> io.BytesIO:
    wb = Workbook()
    sheet_harga(wb.active, prices, month)
    wb.active.title = 'Harga Harian'
    return as_bytes(wb)


def wb_full(txs: list[dict], fin: dict, prices: list[dict], periode: str) -> io.BytesIO:
    """Satu file, banyak sheet: transaksi, rekap pemilik (+per orang), rekap harian, laba rugi, harga."""
    wb = Workbook()
    used: set[str] = set()
    ws = wb.active
    ws.title = safe_sheet('Buku Penerimaan', used)
    sheet_transaksi(ws, txs, periode)
    sheet_rekap_pemilik(wb.create_sheet(safe_sheet('Rekap Pemilik', used)), txs, periode)
    for o in sorted(_owner_rows(txs).values(), key=lambda x: -x['netto']):
        sheet_pemilik_detail(wb.create_sheet(safe_sheet(o['name'], used)), o, periode)
    sheet_rekap_harian(wb.create_sheet(safe_sheet('Rekap Harian', used)), txs, periode)
    sheet_laba_rugi(wb.create_sheet(safe_sheet('Laba Rugi', used)), fin, periode)
    sheet_detail_finance(wb.create_sheet(safe_sheet('Detail Grading', used)), fin, periode)
    sheet_harga(wb.create_sheet(safe_sheet('Harga Harian', used)), prices, None)
    return as_bytes(wb)
