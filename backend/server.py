from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path(__file__).parent / '.env')

import asyncio, io, logging, os, random, secrets, uuid
from datetime import datetime, timezone, timedelta, time as dtime
from typing import Optional
import bcrypt, jwt, pandas as pd
from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, Depends
from fastapi.responses import StreamingResponse
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field
from pymongo import ReturnDocument
from starlette.middleware.cors import CORSMiddleware

logging.basicConfig(level=logging.INFO)
log = logging.getLogger('sjm')
client = AsyncIOMotorClient(os.environ['MONGO_URL'])
db = client[os.environ['DB_NAME']]
app = FastAPI(title='PT SJM Sawit')
api = APIRouter(prefix='/api')
ADMIN_EMAIL = os.environ['ADMIN_EMAIL'].strip().lower()
ADMIN_PASSWORD = os.environ['ADMIN_PASSWORD']
JWT_SECRET = os.environ['JWT_SECRET']
ADMIN_NOTIFY = ADMIN_EMAIL
RESEND_API_KEY = os.environ.get('RESEND_API_KEY', '')

def now(): return datetime.now(timezone.utc).isoformat()
def hash_pw(value): return bcrypt.hashpw(value.encode(), bcrypt.gensalt()).decode()
def check_pw(value, hashed): return bcrypt.checkpw(value.encode(), hashed.encode())
def token(user): return jwt.encode({'sub': user['id'], 'exp': datetime.now(timezone.utc)+timedelta(days=1)}, JWT_SECRET, algorithm='HS256')
def public_user(u): return {k:u.get(k) for k in ('id','name','email','role','status')}

async def notify(subject, html):
    if not RESEND_API_KEY:
        log.info('[MOCKED EMAIL via Resend] to=%s subject=%s', ADMIN_NOTIFY, subject)
        return 'mocked'
    def send():
        import resend
        resend.api_key = RESEND_API_KEY
        resend.Emails.send({'from':'PT SJM <onboarding@resend.dev>','to':[ADMIN_NOTIFY],'subject':subject,'html':html})
    try:
        await asyncio.to_thread(send)
        return 'sent'
    except Exception as exc:
        log.warning('Resend failed: %s', exc); return 'error'

async def next_code_seq(batch_id, prefix):
    docs=await db.transactions.find({'batch_id':batch_id},{'_id':0,'transaction_code':1}).to_list(10000)
    mx=0
    for d in docs:
        c=d.get('transaction_code') or ''
        if c.startswith(prefix):
            try: mx=max(mx,int(c.rsplit('-',1)[1]))
            except ValueError: pass
    return mx+1

MONTHS_ID=['Januari','Februari','Maret','April','Mei','Juni','Juli','Agustus','September','Oktober','November','Desember']
def month_label(m):
    y,mm=m.split('-'); return f'{MONTHS_ID[int(mm)-1]} {y}'

async def current_user(request: Request):
    raw = request.headers.get('Authorization','')
    if not raw.startswith('Bearer '): raise HTTPException(401, 'Sesi login tidak ditemukan.')
    try: payload = jwt.decode(raw[7:], JWT_SECRET, algorithms=['HS256'])
    except jwt.PyJWTError: raise HTTPException(401, 'Sesi login sudah berakhir.')
    user = await db.users.find_one({'id': payload['sub']}, {'_id':0})
    if not user or user.get('status') != 'approved': raise HTTPException(403, 'Akun belum disetujui admin.')
    return user
async def admin_user(user=Depends(current_user)):
    if user.get('role') != 'admin': raise HTTPException(403, 'Akses admin diperlukan.')
    return user

class Signup(BaseModel): name: str; email: str; password: str = Field(min_length=6)
class Login(BaseModel): email: str; password: str
class Owner(BaseModel): id: Optional[str]=None; name: str; vehicle_type: str; capacity_kg: int; active: bool=True
class Price(BaseModel): date: str; price_per_kg: int
class Tx(BaseModel):
    id: Optional[str]=None; batch_id: str; date: str; arrival_time: str; exit_time: str
    owner_id: str; owner_name: str; vehicle_type: str; gross_kg: int; tare_kg: Optional[int]=None
    price_per_kg: int; note: str=''; transaction_code: Optional[str]=None
class Generate(BaseModel): start_date: str; end_date: str; target_kg: int; price_per_kg: int=3000
class NewBatch(BaseModel): month: Optional[str]=None
class FinanceDay(BaseModel): date: str; total_tare_kg: int; sip_price_per_kg: int; freight_per_kg: int=200; batch_id: Optional[str]=None

@app.on_event('startup')
async def startup():
    await db.users.create_index('email', unique=True)
    await db.login_attempts.create_index('identifier')
    admin = await db.users.find_one({'email':ADMIN_EMAIL})
    if not admin:
        await db.users.insert_one({'id':str(uuid.uuid4()),'name':'Admin Utama PT SJM','email':ADMIN_EMAIL,'password_hash':hash_pw(ADMIN_PASSWORD),'role':'admin','status':'approved','created_at':now()})
    elif not check_pw(ADMIN_PASSWORD, admin.get('password_hash','')):
        await db.users.update_one({'email':ADMIN_EMAIL},{'$set':{'password_hash':hash_pw(ADMIN_PASSWORD),'role':'admin','status':'approved'}})
    active = await db.batches.find_one({'active':True})
    if not active:
        m=datetime.now().strftime('%Y-%m')
        await db.batches.insert_one({'id':f'BATCH-{m}', 'label':month_label(m), 'month':m, 'active':True, 'created_at':now()})
    if await db.owners.count_documents({}) == 0:
        await db.owners.insert_many([{'id':str(uuid.uuid4()),'name':'Ita Sari','vehicle_type':'Pick Up','capacity_kg':1300,'active':True},{'id':str(uuid.uuid4()),'name':'Epit','vehicle_type':'Tossa','capacity_kg':600,'active':True},{'id':str(uuid.uuid4()),'name':'Sita Rosiani','vehicle_type':'Hilux','capacity_kg':1800,'active':True},{'id':str(uuid.uuid4()),'name':'Suparmin','vehicle_type':'Hilux','capacity_kg':1800,'active':True}])

@api.post('/auth/signup')
async def signup(p: Signup):
    email=p.email.strip().lower()
    if await db.users.find_one({'email':email}): raise HTTPException(400,'Email sudah terdaftar.')
    doc={'id':str(uuid.uuid4()),'name':p.name.strip(),'email':email,'password_hash':hash_pw(p.password),'role':'operator','status':'pending','created_at':now()}
    await db.users.insert_one(doc); await notify('Signup baru menunggu approval PT SJM',f'<p>{p.name} ({email}) menunggu persetujuan.</p>')
    return {'message':'Pendaftaran berhasil. Tunggu persetujuan admin sebelum login.','status':'pending'}

@api.post('/auth/login')
async def login(p: Login, request: Request):
    email=p.email.strip().lower()
    ip=request.headers.get('x-forwarded-for','').split(',')[0].strip() or (request.client.host if request.client else 'unknown')
    ident=f'{ip}:{email}'
    att=await db.login_attempts.find_one({'identifier':ident})
    if att and att.get('count',0)>=5:
        if datetime.now(timezone.utc)<datetime.fromisoformat(att.get('locked_until') or now()):
            raise HTTPException(423,'Terlalu banyak percobaan gagal. Akun dikunci sementara, coba lagi dalam 15 menit.')
        await db.login_attempts.delete_one({'identifier':ident})
    u=await db.users.find_one({'email':email})
    if not u or not check_pw(p.password,u.get('password_hash','')):
        r=await db.login_attempts.find_one_and_update({'identifier':ident},{'$inc':{'count':1}},upsert=True,return_document=ReturnDocument.AFTER)
        if r['count']>=5:
            await db.login_attempts.update_one({'identifier':ident},{'$set':{'locked_until':(datetime.now(timezone.utc)+timedelta(minutes=15)).isoformat()}})
        raise HTTPException(401,'Email atau password salah.')
    if u.get('status')!='approved': raise HTTPException(403,'Akun masih menunggu approval admin.')
    await db.login_attempts.delete_one({'identifier':ident})
    await notify('Notifikasi login PT SJM',f'<p>Login {u["name"]} ({email}) terdeteksi pada {now()}.</p>')
    return {'user':public_user(u),'token':token(u)}

@api.get('/auth/me')
async def me(user=Depends(current_user)): return user
@api.get('/batches')
async def batches(user=Depends(current_user)): return await db.batches.find({}, {'_id':0}).sort('created_at',-1).to_list(100)
@api.post('/batches/new')
async def new_batch(p: Optional[NewBatch]=None, user=Depends(current_user)):
    month=(p.month.strip() if p and p.month else datetime.now().strftime('%Y-%m'))
    try: datetime.strptime(month,'%Y-%m')
    except ValueError: raise HTTPException(400,'Format bulan harus YYYY-MM.')
    await db.batches.update_many({'active':True},{'$set':{'active':False,'archived_at':now()}})
    n=await db.batches.count_documents({'month':month})
    label=month_label(month)+(f' · Periode {n+1}' if n else '')
    bid=f'BATCH-{month}' if n==0 else f'BATCH-{month}-{n+1}'
    doc={'id':bid,'label':label,'month':month,'active':True,'created_at':now()}
    await db.batches.insert_one(doc); return {k:v for k,v in doc.items() if k!='_id'}

@api.post('/batches/{bid}/activate')
async def activate_batch(bid, user=Depends(current_user)):
    b=await db.batches.find_one({'id':bid},{'_id':0})
    if not b: raise HTTPException(404,'Batch tidak ditemukan.')
    await db.batches.update_many({'active':True},{'$set':{'active':False,'archived_at':now()}})
    await db.batches.update_one({'id':bid},{'$set':{'active':True},'$unset':{'archived_at':''}})
    return {'success':True,'id':bid}

@api.delete('/batches/{bid}')
async def delete_batch(bid, user=Depends(admin_user)):
    b=await db.batches.find_one({'id':bid})
    if not b: raise HTTPException(404,'Batch tidak ditemukan.')
    txd=(await db.transactions.delete_many({'batch_id':bid})).deleted_count
    await db.batches.delete_one({'id':bid})
    if b.get('active'):
        latest=await db.batches.find({}).sort('created_at',-1).to_list(1)
        if latest: await db.batches.update_one({'id':latest[0]['id']},{'$set':{'active':True},'$unset':{'archived_at':''}})
        else:
            m=datetime.now().strftime('%Y-%m')
            await db.batches.insert_one({'id':f'BATCH-{m}','label':month_label(m),'month':m,'active':True,'created_at':now()})
    await notify('Batch dihapus PT SJM',f"<p>Batch {b.get('label',bid)} dihapus beserta {txd} transaksi.</p>")
    return {'success':True,'deleted_transactions':txd}
@api.get('/owners')
async def owners(user=Depends(current_user)): return await db.owners.find({}, {'_id':0}).to_list(100)
@api.post('/owners')
async def save_owner(p: Owner, user=Depends(current_user)):
    d=p.model_dump(); d['id']=d.get('id') or str(uuid.uuid4()); await db.owners.update_one({'id':d['id']},{'$set':d},upsert=True); return d
@api.get('/prices')
async def prices(user=Depends(current_user)): return await db.prices.find({}, {'_id':0}).sort('date',-1).to_list(400)
@api.post('/prices')
async def save_price(p: Price, user=Depends(current_user)): d=p.model_dump(); await db.prices.update_one({'date':p.date},{'$set':d},upsert=True); return d

def validate_tx(d, owners):
    owner=next((o for o in owners if o['id']==d['owner_id']),None)
    if not owner: raise HTTPException(400,'Pemilik tidak ditemukan.')
    if d['gross_kg']>owner['capacity_kg']: raise HTTPException(400,f"Gross melebihi kapasitas {owner['vehicle_type']} ({owner['capacity_kg']} Kg).")
    arr=datetime.strptime(f"{d['date']} {d['arrival_time']}",'%Y-%m-%d %H:%M:%S'); ex=datetime.strptime(f"{d['date']} {d['exit_time']}",'%Y-%m-%d %H:%M:%S')
    if not (datetime.strptime('07:00:00','%H:%M:%S').time()<=arr.time()<=datetime.strptime('16:00:00','%H:%M:%S').time()): raise HTTPException(400,'Jam terima harus 07:00:00–16:00:00.')
    if datetime.strptime('12:00:00','%H:%M:%S').time()<=arr.time()<datetime.strptime('13:00:00','%H:%M:%S').time(): raise HTTPException(400,'Tidak ada penerimaan pada jam istirahat 12:00–13:00.')
    if ex<=arr: raise HTTPException(400,'Jam keluar harus setelah jam terima.')
    d['owner_name']=owner['name']; d['vehicle_type']=owner['vehicle_type']; return d

@api.get('/transactions')
async def transactions(batch_id:Optional[str]=None,user=Depends(current_user)):
    if not batch_id:
        b=await db.batches.find_one({'active':True}); batch_id=b['id'] if b else None
    return await db.transactions.find({'batch_id':batch_id},{'_id':0}).sort([('date',-1),('arrival_time',-1)]).to_list(5000)
@api.post('/transactions')
async def save_tx(p:Tx,user=Depends(current_user)):
    d=p.model_dump(); owners=await db.owners.find({}, {'_id':0}).to_list(100); d=validate_tx(d,owners); d['id']=d.get('id') or str(uuid.uuid4()); d['netto_kg']=d['gross_kg']-(d.get('tare_kg') or 0); d['netto_ton']=round(d['netto_kg']/1000,2); d['total_amount']=d['netto_kg']*d['price_per_kg']
    if not d.get('transaction_code'):
        pref=f"SJM-{d['date'][:4]}{d['date'][5:7]}-"; d['transaction_code']=f"{pref}{await next_code_seq(d['batch_id'],pref):03d}"
    await db.transactions.update_one({'id':d['id']},{'$set':d},upsert=True); await notify('Edit transaksi PT SJM',f"<p>Transaksi {d['transaction_code']} diedit/disimpan.</p>"); return {**d,'_id':None}
@api.delete('/transactions/{tx_id}')
async def delete_tx(tx_id,user=Depends(current_user)): await db.transactions.delete_one({'id':tx_id}); return {'success':True}
@api.delete('/transactions/date/{date}')
async def delete_day(date,batch_id:Optional[str]=None,user=Depends(current_user)):
    if not batch_id:
        b=await db.batches.find_one({'active':True}); batch_id=b['id'] if b else None
    if not batch_id: raise HTTPException(400,'Batch tidak ditemukan.')
    r=await db.transactions.delete_many({'date':date,'batch_id':batch_id})
    await notify('Hapus transaksi harian PT SJM',f'<p>{r.deleted_count} transaksi tanggal {date} dihapus dari batch {batch_id}.</p>')
    return {'deleted_count':r.deleted_count,'date':date,'batch_id':batch_id}

def build_schedule(days, owners, target_kg, allow_lunch):
    candidates=[]; remaining=target_kg
    for day in days:
        if remaining<=0: break
        t=datetime(day.year,day.month,day.day,7,0,0)+timedelta(minutes=random.randint(2,28),seconds=random.randint(0,59))
        prev_owner=None; last_arr={}
        while remaining>0 and t.time()<=dtime(16,0):
            if not allow_lunch and dtime(12,0)<=t.time()<dtime(13,0):
                t=datetime(day.year,day.month,day.day,13,0,0)+timedelta(minutes=random.randint(1,9),seconds=random.randint(0,59)); continue
            avail=[o for o in owners if o['id']!=prev_owner and (o['id'] not in last_arr or (t-last_arr[o['id']]).total_seconds()>=105*60)]
            if not avail:
                t+=timedelta(minutes=random.randint(8,15),seconds=random.randint(0,59)); continue
            avail.sort(key=lambda o:last_arr.get(o['id'],datetime(2000,1,1)))
            o=random.choice(avail[:2])
            kg=min(o['capacity_kg']-random.randint(20,120),remaining)
            ex=t+timedelta(minutes=random.randint(20,40),seconds=random.randint(5,55))
            lunch=dtime(12,0)<=t.time()<dtime(13,0)
            note='Penerimaan tambahan melewati jam istirahat untuk mengejar target.' if lunch else ('Muatan dikurangi menyesuaikan sisa target.' if kg<o['capacity_kg']*.6 else 'Muatan normal sesuai kapasitas kendaraan.')
            candidates.append({'date':day.strftime('%Y-%m-%d'),'arrival_time':t.strftime('%H:%M:%S'),'exit_time':ex.strftime('%H:%M:%S'),'owner_id':o['id'],'owner_name':o['name'],'vehicle_type':o['vehicle_type'],'gross_kg':kg,'tare_kg':None,'note':note})
            remaining-=kg; prev_owner=o['id']; last_arr[o['id']]=t
            t=ex+timedelta(minutes=random.randint(4,14),seconds=random.randint(0,59))
    return candidates,remaining

@api.post('/generator/run')
async def generate(p:Generate,user=Depends(current_user)):
    b=await db.batches.find_one({'active':True}); owners=await db.owners.find({'active':True},{'_id':0}).to_list(20)
    if not b or not owners: raise HTTPException(400,'Batch atau master pemilik belum tersedia.')
    start=datetime.strptime(p.start_date,'%Y-%m-%d'); end=datetime.strptime(p.end_date,'%Y-%m-%d')
    if end<start: raise HTTPException(400,'Tanggal akhir harus sama atau setelah tanggal mulai.')
    if p.target_kg<=0: raise HTTPException(400,'Target harus lebih dari 0 Kg.')
    days=[start+timedelta(days=i) for i in range((end-start).days+1)]
    cands,rem=[],p.target_kg
    for _ in range(6):
        c,r=build_schedule(days,owners,p.target_kg,False)
        if r<rem: cands,rem=c,r
        if rem<=0: break
    used_lunch=False
    if rem>0:
        for _ in range(6):
            c,r=build_schedule(days,owners,p.target_kg,True)
            if r<rem: cands,rem,used_lunch=c,r,True
            if rem<=0: break
    if rem>0:
        raise HTTPException(422,detail={'message':'Target tidak dapat dicapai dengan konfigurasi kendaraan dan jam operasional saat ini. Tidak ada transaksi yang disimpan.','target':p.target_kg,'maximum':p.target_kg-rem,'shortage':rem,'reason':'Slot operasional dan kapasitas kendaraan tidak mencukupi meski melewati jam istirahat.','status':'infeasible'})
    seqs={}
    for d in cands:
        pref=f"SJM-{d['date'][:4]}{d['date'][5:7]}-"
        if pref not in seqs: seqs[pref]=await next_code_seq(b['id'],pref)
        d.update({'id':str(uuid.uuid4()),'batch_id':b['id'],'price_per_kg':p.price_per_kg,'netto_kg':d['gross_kg'],'netto_ton':round(d['gross_kg']/1000,2),'total_amount':d['gross_kg']*p.price_per_kg,'transaction_code':f'{pref}{seqs[pref]:03d}'}); seqs[pref]+=1
    await db.transactions.insert_many([dict(d) for d in cands])
    await notify('Generator PT SJM selesai',f"<p>{len(cands)} transaksi dibuat, total {p.target_kg} Kg ({p.start_date} s/d {p.end_date}).</p>")
    return {'generated_count':len(cands),'target_kg':p.target_kg,'actual_kg':sum(d['gross_kg'] for d in cands),'used_lunch_break':used_lunch,'status':'exact'}

@api.post('/finance/day')
async def finance_day(p:FinanceDay,user=Depends(current_user)):
    batch_id=p.batch_id
    if not batch_id:
        b=await db.batches.find_one({'active':True}); batch_id=b['id'] if b else None
    txs=await db.transactions.find({'batch_id':batch_id,'date':p.date},{'_id':0}).sort('arrival_time',1).to_list(1000)
    if not txs: raise HTTPException(400,'Tidak ada transaksi pada tanggal tersebut.')
    total_netto=sum(t['netto_kg'] for t in txs)
    if not (0<=p.total_tare_kg<total_netto): raise HTTPException(400,f'Total tare harus antara 0 dan kurang dari total netto hari itu ({total_netto} Kg).')
    if p.sip_price_per_kg<=0 or p.freight_per_kg<0: raise HTTPException(400,'Harga PT SIP dan biaya angkut tidak valid.')
    raw=[p.total_tare_kg*t['netto_kg']/total_netto*random.uniform(.75,1.25) for t in txs]
    tares=[max(0,min(int(round(r)),t['netto_kg']-1)) for r,t in zip(raw,txs)]
    diff=p.total_tare_kg-sum(tares); i=0
    while diff!=0 and i<200000:
        j=i%len(txs); step=1 if diff>0 else -1
        if 0<=tares[j]+step<=txs[j]['netto_kg']-1: tares[j]+=step; diff-=step
        i+=1
    if diff!=0: raise HTTPException(400,'Distribusi tare gagal. Periksa nilai total tare.')
    for t,tr in zip(txs,tares):
        await db.transactions.update_one({'id':t['id']},{'$set':{'grading_tare_kg':tr}})
    doc={'batch_id':batch_id,'date':p.date,'total_tare_kg':p.total_tare_kg,'sip_price_per_kg':p.sip_price_per_kg,'freight_per_kg':p.freight_per_kg,'updated_at':now()}
    await db.finance_days.update_one({'batch_id':batch_id,'date':p.date},{'$set':doc},upsert=True)
    return {'success':True,'date':p.date,'total_tare_kg':p.total_tare_kg,'distributed':len(txs)}

@api.get('/finance/summary')
async def finance_summary(batch_id:Optional[str]=None,start:Optional[str]=None,end:Optional[str]=None,user=Depends(current_user)):
    if not batch_id:
        b=await db.batches.find_one({'active':True}); batch_id=b['id'] if b else None
    q={'batch_id':batch_id}
    if start or end: q['date']={}; q['date'].update({'$gte':start} if start else {}); q['date'].update({'$lte':end} if end else {})
    txs=await db.transactions.find(q,{'_id':0}).to_list(10000)
    fdays={f['date']:f for f in await db.finance_days.find({'batch_id':batch_id},{'_id':0}).to_list(1000)}
    days={}
    for t in txs: days.setdefault(t['date'],[]).append(t)
    rows=[]; detail=[]
    for date in sorted(days):
        dtx=sorted(days[date],key=lambda x:x['arrival_time']); f=fdays.get(date)
        netto=sum(t['netto_kg'] for t in dtx); beli=sum(t['total_amount'] for t in dtx)
        rate=f['freight_per_kg'] if f else 200; angkut=netto*rate; modal=beli+angkut
        row={'date':date,'tx_count':len(dtx),'netto_kg':netto,'total_beli':beli,'freight_per_kg':rate,'total_angkut':angkut,'total_modal':modal,'configured':bool(f)}
        if f:
            tare=f['total_tare_kg']; njual=netto-tare; hjual=njual*f['sip_price_per_kg']; pph=round(hjual*0.0025); untung=hjual-pph-modal
            row.update({'total_tare_kg':tare,'grading_pct':round(tare/netto*100,2),'netto_jual':njual,'sip_price_per_kg':f['sip_price_per_kg'],'harga_jual':hjual,'pph22':pph,'untung_rugi':untung})
        rows.append(row)
        for t in dtx:
            g=t.get('grading_tare_kg')
            d={'id':t['id'],'code':t['transaction_code'],'date':date,'owner_name':t['owner_name'],'netto_kg':t['netto_kg'],'price_per_kg':t['price_per_kg'],'total_amount':t['total_amount'],'freight':t['netto_kg']*rate}
            if f and g is not None:
                d.update({'tare_kg':g,'grading_pct':round(g/t['netto_kg']*100,2),'netto_jual':t['netto_kg']-g,'harga_jual':(t['netto_kg']-g)*f['sip_price_per_kg']})
            detail.append(d)
    conf=[r for r in rows if r['configured']]
    totals={'netto_kg':sum(r['netto_kg'] for r in rows),'total_beli':sum(r['total_beli'] for r in rows),'total_angkut':sum(r['total_angkut'] for r in rows),'total_modal':sum(r['total_modal'] for r in rows),'total_tare_kg':sum(r.get('total_tare_kg',0) for r in conf),'netto_jual':sum(r.get('netto_jual',0) for r in conf),'harga_jual':sum(r.get('harga_jual',0) for r in conf),'pph22':sum(r.get('pph22',0) for r in conf),'untung_rugi':sum(r.get('untung_rugi',0) for r in conf),'configured_days':len(conf),'total_days':len(rows)}
    return {'days':rows,'totals':totals,'transactions':detail}

@api.get('/analytics/summary')
async def summary(batch_id:Optional[str]=None,start:Optional[str]=None,end:Optional[str]=None,user=Depends(current_user)):
    if not batch_id:
        b=await db.batches.find_one({'active':True}); batch_id=b['id'] if b else None
    q={'batch_id':batch_id}
    if start or end: q['date']={}; q['date'].update({'$gte':start} if start else {}); q['date'].update({'$lte':end} if end else {})
    txs=await db.transactions.find(q,{'_id':0}).to_list(10000); daily={}; by_owner={}
    for t in txs:
        for target,key in ((daily,t['date']),(by_owner,t['owner_name'])): target.setdefault(key,{'label':key,'count':0,'netto_kg':0,'total_amount':0}); target[key]['count']+=1; target[key]['netto_kg']+=t['netto_kg']; target[key]['total_amount']+=t['total_amount']
    return {'total_transactions':len(txs),'total_netto_kg':sum(t['netto_kg'] for t in txs),'total_netto_ton':round(sum(t['netto_kg'] for t in txs)/1000,2),'total_spending':sum(t['total_amount'] for t in txs),'daily_summary':list(daily.values()),'owner_summary':list(by_owner.values()),'transactions':txs}
@api.get('/export/excel')
async def export(batch_id:Optional[str]=None,start:Optional[str]=None,end:Optional[str]=None,user=Depends(current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    s=await summary(batch_id,start,end,user=user)
    wb=Workbook(); ws=wb.active; ws.title='Rekap PT SJM'
    heads=['Kode','Tanggal','Jam Terima','Jam Keluar','Pemilik','Kendaraan','Gross Kg','Tare Kg','Netto Kg','Netto Ton','Harga/Kg','Total Beli']
    ws.append(heads)
    for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='15803D'); c.alignment=Alignment(horizontal='center')
    txs=sorted(s['transactions'],key=lambda t:(t['date'],t['arrival_time']))
    for r,t in enumerate(txs,start=2):
        ws.append([t['transaction_code'],t['date'],t['arrival_time'],t['exit_time'],t['owner_name'],t['vehicle_type'],t['gross_kg'],t.get('tare_kg') or 0,f'=G{r}-H{r}',f'=ROUND((G{r}-H{r})/1000,2)',t['price_per_kg'],f'=(G{r}-H{r})*K{r}'])
    last=len(txs)+1
    if txs:
        ws.append(['TOTAL','','','','','',f'=SUM(G2:G{last})',f'=SUM(H2:H{last})',f'=SUM(I2:I{last})',f'=SUM(J2:J{last})','',f'=SUM(L2:L{last})'])
        for c in ws[last+1]: c.font=Font(bold=True); c.fill=PatternFill('solid',fgColor='E9F5ED')
    for i,w in enumerate([16,12,11,11,17,12,10,9,10,11,11,15],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':'attachment; filename=Rekap_PT_SJM.xlsx'})
@api.get('/admin/users')
async def admin_users(user=Depends(admin_user)): return await db.users.find({}, {'_id':0,'password_hash':0}).to_list(1000)
@api.post('/admin/users/{uid}/approve')
async def approve(uid,user=Depends(admin_user)): r=await db.users.update_one({'id':uid},{'$set':{'status':'approved'}}); return {'success':r.modified_count>0}
@api.delete('/admin/users/{uid}')
async def remove_user(uid,user=Depends(admin_user)): await db.users.delete_one({'id':uid,'email':{'$ne':ADMIN_EMAIL}}); return {'success':True}
app.include_router(api)
app.add_middleware(CORSMiddleware,allow_origins=[o.strip() for o in os.environ.get('CORS_ORIGINS','*').split(',')],allow_credentials=True,allow_methods=['*'],allow_headers=['*'])
@app.on_event('shutdown')
async def shutdown(): client.close()