"""Phase 3 comprehensive testing"""
import requests
from io import BytesIO
from openpyxl import load_workbook

BASE_URL = "https://commit-history-web.preview.emergentagent.com"

print("\n" + "="*60)
print("PHASE 3 BACKEND TESTING")
print("="*60)

# Test 1: Admin Accounts
print("\n[1/4] Testing Admin Accounts...")
r1 = requests.post(f"{BASE_URL}/api/auth/login", 
                   json={"email": "yumaclovstar@gmail.com", "password": "178910"}, 
                   timeout=30)
assert r1.status_code == 200, f"Primary admin login failed"
token1 = r1.json()["token"]
print(f"  ✅ Primary admin: yumaclovstar@gmail.com (role={r1.json()['user']['role']})")

r2 = requests.post(f"{BASE_URL}/api/auth/login",
                   json={"email": "admin@ptsjm.co.id", "password": "AdminSJM2026!"},
                   timeout=30)
assert r2.status_code == 200, f"Backup admin login failed"
print(f"  ✅ Backup admin: admin@ptsjm.co.id (role={r2.json()['user']['role']})")

# Test 2: Prices CRUD
print("\n[2/4] Testing Prices CRUD...")
headers = {"Authorization": f"Bearer {token1}"}

# Create/update prices
test_date = "2026-09-10"
r = requests.post(f"{BASE_URL}/api/prices",
                 json={"date": test_date, "price_per_kg": 3100, "note": "Test price"},
                 headers=headers, timeout=30)
assert r.status_code == 200
print(f"  ✅ Created/updated price for {test_date}")

# Get prices
r = requests.get(f"{BASE_URL}/api/prices?month=2026-09", headers=headers, timeout=30)
assert r.status_code == 200
data = r.json()
print(f"  ✅ Retrieved {len(data['prices'])} prices for 2026-09")
print(f"     Stats: high={data['stats']['high']}, low={data['stats']['low']}, avg={data['stats']['avg']}")

# Verify our test price exists
prices_dict = {p["date"]: p for p in data["prices"]}
assert test_date in prices_dict
assert prices_dict[test_date]["price_per_kg"] == 3100
print(f"  ✅ Verified test price exists and is correct")

# Delete test price
r = requests.delete(f"{BASE_URL}/api/prices/{test_date}", headers=headers, timeout=30)
assert r.status_code == 200
print(f"  ✅ Deleted test price")

# Test 3: Generate test data for exports
print("\n[3/4] Generating test data for exports...")
r = requests.get(f"{BASE_URL}/api/batches", headers=headers, timeout=30)
batches = r.json()
batch_id = next(b["id"] for b in batches if b["active"])
print(f"  Active batch: {batch_id}")

# Run generator to create transactions
r = requests.post(f"{BASE_URL}/api/generator/run",
                 json={"start_date": "2026-09-15", "end_date": "2026-09-15",
                       "target_kg": 5000, "price_per_kg": 2900},
                 headers=headers, timeout=120)
if r.status_code == 200:
    result = r.json()
    print(f"  ✅ Generated {result['generated_count']} transactions ({result['actual_kg']} Kg)")
else:
    print(f"  ⚠️  Generator returned {r.status_code} (may already have data)")

# Test 4: Excel Exports
print("\n[4/4] Testing Excel Exports...")

# Complete export
r = requests.get(f"{BASE_URL}/api/export/excel?batch_id={batch_id}", 
                 headers=headers, timeout=60)
assert r.status_code == 200
assert "spreadsheetml" in r.headers.get("content-type", "")
wb = load_workbook(BytesIO(r.content))
expected = ["Buku Penerimaan", "Rekap Pemilik", "Rekap Harian", "Laba Rugi", "Harga Harian"]
for sheet in expected:
    assert sheet in wb.sheetnames, f"Missing sheet: {sheet}"
print(f"  ✅ Complete export: {len(wb.sheetnames)} sheets")

# Owners export (all)
r = requests.get(f"{BASE_URL}/api/export/owners-excel?batch_id={batch_id}",
                 headers=headers, timeout=60)
assert r.status_code == 200
wb = load_workbook(BytesIO(r.content))
assert "Ringkasan Pemilik" in wb.sheetnames
print(f"  ✅ Owners export (all): {len(wb.sheetnames)} sheets")

# Owners export (single) - get first owner with transactions
r = requests.get(f"{BASE_URL}/api/transactions?batch_id={batch_id}", headers=headers, timeout=30)
txs = r.json()
if txs:
    owner_name = txs[0]["owner_name"]
    r = requests.get(f"{BASE_URL}/api/export/owners-excel?batch_id={batch_id}&owner={owner_name}",
                     headers=headers, timeout=60)
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert len(wb.sheetnames) == 1
    print(f"  ✅ Owners export (single): {owner_name}")
else:
    print(f"  ⚠️  No transactions to test single owner export")

# Daily export
r = requests.get(f"{BASE_URL}/api/export/daily-excel?batch_id={batch_id}",
                 headers=headers, timeout=60)
assert r.status_code == 200
wb = load_workbook(BytesIO(r.content))
assert "Rekap Harian" in wb.sheetnames
print(f"  ✅ Daily export: {wb.sheetnames}")

# Finance export
r = requests.get(f"{BASE_URL}/api/export/finance-excel?batch_id={batch_id}",
                 headers=headers, timeout=60)
assert r.status_code == 200
wb = load_workbook(BytesIO(r.content))
assert "Laba Rugi Harian" in wb.sheetnames
print(f"  ✅ Finance export: {wb.sheetnames}")

# Prices export
r = requests.get(f"{BASE_URL}/api/export/prices-excel?month=2026-09",
                 headers=headers, timeout=60)
assert r.status_code == 200
wb = load_workbook(BytesIO(r.content))
assert "Harga Harian" in wb.sheetnames
print(f"  ✅ Prices export: Harga Harian sheet present")

print("\n" + "="*60)
print("✅ ALL BACKEND TESTS PASSED")
print("="*60 + "\n")
