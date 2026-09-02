"""PPh 22 yang bisa dikelola: tarif default global, tarif khusus per tanggal, nominal manual, reset.

Semua pekerjaan destruktif memakai batch buangan (2026-06) yang dihapus di akhir,
dan tarif default dikembalikan ke 0.25% setelah selesai.
"""
import io

import pytest
import requests
from openpyxl import load_workbook

from conftest import BASE_URL

MONTH = "2026-06"
D1 = "2026-06-02"


@pytest.fixture(scope="module")
def admin(test_credentials):
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    r = s.post(f"{BASE_URL}/api/auth/login", json=test_credentials, timeout=60)
    assert r.status_code == 200, r.text[:200]
    s.headers.update({"Authorization": f"Bearer {r.json()['token']}"})
    return s


@pytest.fixture(scope="module")
def graded_batch(admin):
    """Batch buangan + 1 hari transaksi yang sudah digrading."""
    prev = [b for b in admin.get(f"{BASE_URL}/api/batches", timeout=60).json() if b.get("active")]
    r = admin.post(f"{BASE_URL}/api/batches/new", json={"month": MONTH}, timeout=60)
    assert r.status_code == 200, r.text
    bid = r.json()["id"]
    g = admin.post(f"{BASE_URL}/api/generator/run", json={
        "start_date": D1, "end_date": D1, "target_kg": 5000, "price_per_kg": 2500}, timeout=120)
    assert g.status_code == 200, g.text
    f = admin.post(f"{BASE_URL}/api/finance/day", json={
        "batch_id": bid, "date": D1, "total_tare_kg": 200,
        "sip_price_per_kg": 3500, "freight_per_kg": 200}, timeout=60)
    assert f.status_code == 200, f.text
    yield bid
    admin.post(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.25}, timeout=30)
    admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.25}, timeout=30)
    admin.delete(f"{BASE_URL}/api/batches/{bid}", timeout=60)
    if prev:
        admin.post(f"{BASE_URL}/api/batches/{prev[0]['id']}/activate", timeout=60)


def row(admin, batch, date=D1):
    r = admin.get(f"{BASE_URL}/api/finance/summary?batch_id={batch}", timeout=60)
    assert r.status_code == 200, r.text
    data = r.json()
    return next(d for d in data["days"] if d["date"] == date), data["totals"]


class TestPphManaged:
    def test_default_rate_is_quarter_percent(self, admin, graded_batch):
        assert admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.25}, timeout=30).status_code == 200
        s = admin.get(f"{BASE_URL}/api/settings/pph", timeout=30).json()
        assert s["rate_pct"] == 0.25 and s["system_default"] == 0.25
        d, tot = row(admin, graded_batch)
        assert d["pph_mode"] == "auto" and d["pph_rate_pct"] == 0.25 and d["pph_custom"] is False
        assert d["pph22"] == round(d["harga_jual"] * 0.0025)
        assert d["untung_rugi"] == d["harga_jual"] - d["pph22"] - d["total_modal"]
        assert tot["pph_rate_default"] == 0.25

    def test_global_default_rate_change_applies(self, admin, graded_batch):
        assert admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.5}, timeout=30).status_code == 200
        d, tot = row(admin, graded_batch)
        assert d["pph_rate_pct"] == 0.5 and d["pph22"] == round(d["harga_jual"] * 0.005)
        assert d["untung_rugi"] == d["harga_jual"] - d["pph22"] - d["total_modal"]
        assert tot["pph22"] == d["pph22"]
        admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.25}, timeout=30)

    def test_per_day_custom_rate(self, admin, graded_batch):
        r = admin.post(f"{BASE_URL}/api/finance/pph", json={
            "batch_id": graded_batch, "date": D1, "mode": "auto",
            "rate_pct": 1.5, "note": "tarif khusus"}, timeout=60)
        assert r.status_code == 200, r.text
        d, _ = row(admin, graded_batch)
        assert d["pph22"] == round(d["harga_jual"] * 0.015)
        assert d["pph_custom"] is True and d["pph_note"] == "tarif khusus"
        # tarif default berubah TIDAK mempengaruhi tanggal dengan tarif khusus
        admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.9}, timeout=30)
        d2, _ = row(admin, graded_batch)
        assert d2["pph_rate_pct"] == 1.5 and d2["pph22"] == d["pph22"]
        admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 0.25}, timeout=30)

    def test_manual_amount_from_final_result(self, admin, graded_batch):
        r = admin.post(f"{BASE_URL}/api/finance/pph", json={
            "batch_id": graded_batch, "date": D1, "mode": "manual",
            "amount": 123456, "note": "bukti potong PT SIP"}, timeout=60)
        assert r.status_code == 200, r.text
        d, tot = row(admin, graded_batch)
        assert d["pph_mode"] == "manual" and d["pph22"] == 123456
        assert d["untung_rugi"] == d["harga_jual"] - 123456 - d["total_modal"]
        assert d["pph_rate_pct"] == round(123456 / d["harga_jual"] * 100, 4)
        assert tot["pph_manual_days"] == 1 and tot["pph22"] == 123456
        assert tot["pph_effective_pct"] == round(123456 / tot["harga_jual"] * 100, 4)

    def test_validations(self, admin, graded_batch):
        d, _ = row(admin, graded_batch)
        cases = [
            ({"mode": "manual", "amount": d["harga_jual"] + 1}, 400),
            ({"mode": "manual", "amount": -1}, 400),
            ({"mode": "auto", "rate_pct": 120}, 400),
            ({"mode": "aneh"}, 400),
        ]
        for body, code in cases:
            r = admin.post(f"{BASE_URL}/api/finance/pph",
                           json={"batch_id": graded_batch, "date": D1, **body}, timeout=60)
            assert r.status_code == code, (body, r.status_code, r.text[:150])
        # tanggal yang belum digrading tidak boleh diatur PPh-nya
        r = admin.post(f"{BASE_URL}/api/finance/pph", json={
            "batch_id": graded_batch, "date": "2030-01-01", "mode": "auto"}, timeout=60)
        assert r.status_code == 400
        assert admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": -1}, timeout=30).status_code == 400
        assert admin.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 101}, timeout=30).status_code == 400

    def test_excel_shows_pph_mode(self, admin, graded_batch):
        r = admin.get(f"{BASE_URL}/api/export/finance-excel?batch_id={graded_batch}", timeout=120)
        assert r.status_code == 200
        ws = load_workbook(io.BytesIO(r.content))["Laba Rugi Harian"]
        hr = next(i for i in range(1, 12) if ws.cell(row=i, column=1).value == "Tanggal")
        heads = [c.value for c in ws[hr]]
        assert "PPh 22" in heads and "Tarif / Mode PPh" in heads, heads
        tarif = ws.cell(row=hr + 1, column=heads.index("Tarif / Mode PPh") + 1).value
        assert "Manual" in str(tarif), tarif

    def test_reset_back_to_default(self, admin, graded_batch):
        r = admin.delete(f"{BASE_URL}/api/finance/pph/{D1}?batch_id={graded_batch}", timeout=60)
        assert r.status_code == 200, r.text
        d, tot = row(admin, graded_batch)
        assert d["pph_mode"] == "auto" and d["pph_custom"] is False
        assert d["pph22"] == round(d["harga_jual"] * 0.0025)
        assert tot["pph_manual_days"] == 0
        assert admin.delete(f"{BASE_URL}/api/finance/pph/2030-01-01?batch_id={graded_batch}", timeout=60).status_code == 404

    def test_operator_cannot_change_default_rate(self, admin, graded_batch):
        anon = requests.Session()
        assert anon.put(f"{BASE_URL}/api/settings/pph", json={"rate_pct": 5}, timeout=30).status_code in (401, 403)
