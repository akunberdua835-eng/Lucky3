"""Verifikasi cepat fitur baru: login admin baru, prices CRUD, dan 5 export Excel."""
import io, requests
from openpyxl import load_workbook

BASE = 'http://localhost:8001'
s = requests.Session()
r = s.post(f'{BASE}/api/auth/login', json={'email': 'yumaclovstar@gmail.com', 'password': '178910'})
print('login new admin', r.status_code, r.json().get('user'))
s.headers['Authorization'] = f"Bearer {r.json()['token']}"

r = s.post(f'{BASE}/api/auth/login', json={'email': 'admin@ptsjm.co.id', 'password': 'AdminSJM2026!'})
print('login backup admin', r.status_code, r.json().get('user', {}).get('role'))

b = s.get(f'{BASE}/api/batches').json()[0]['id']
print('batch', b)

if not s.get(f'{BASE}/api/transactions?batch_id={b}').json():
    g = s.post(f'{BASE}/api/generator/run', json={'start_date': '2026-09-01', 'end_date': '2026-09-03', 'target_kg': 12000, 'price_per_kg': 2900})
    print('generator', g.status_code, g.json())

txs = s.get(f'{BASE}/api/transactions?batch_id={b}').json()
d0 = sorted({t['date'] for t in txs})[0]
f = s.post(f'{BASE}/api/finance/day', json={'batch_id': b, 'date': d0, 'total_tare_kg': 300, 'sip_price_per_kg': 3500, 'freight_per_kg': 200})
print('finance day', f.status_code, f.json())

for i, (date, price, note) in enumerate([('2026-09-01', 2800, 'Awal bulan'), ('2026-09-02', 2900, 'Naik permintaan pabrik'), ('2026-09-05', 2750, 'Turun, buah mentah banyak')]):
    print('price', s.post(f'{BASE}/api/prices', json={'date': date, 'price_per_kg': price, 'note': note}).status_code, end=' ')
print()
pr = s.get(f'{BASE}/api/prices?month=2026-09').json()
print('prices rows', [(p['date'], p['price_per_kg'], p['change'], p['change_pct'], p['trend']) for p in pr['prices']])
print('stats', pr['stats'])
print('delete price', s.delete(f'{BASE}/api/prices/2026-09-05').status_code)
s.post(f'{BASE}/api/prices', json={'date': '2026-09-05', 'price_per_kg': 2750, 'note': 'Turun, buah mentah banyak'})

owner = txs[0]['owner_name']
targets = [
    ('/api/export/excel', {'batch_id': b}),
    ('/api/export/owners-excel', {'batch_id': b}),
    ('/api/export/owners-excel', {'batch_id': b, 'owner': owner}),
    ('/api/export/daily-excel', {'batch_id': b}),
    ('/api/export/finance-excel', {'batch_id': b}),
    ('/api/export/prices-excel', {'month': '2026-09'}),
]
for path, params in targets:
    r = s.get(f'{BASE}{path}', params=params)
    ok = r.status_code == 200 and r.headers.get('content-type', '').startswith('application/vnd')
    wb = load_workbook(io.BytesIO(r.content)) if ok else None
    print(f"{path} {params.get('owner','')} -> {r.status_code} size={len(r.content)} sheets={wb.sheetnames if wb else r.text[:200]}")
print('DONE')
