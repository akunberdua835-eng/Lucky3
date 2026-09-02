import time
import uuid
from datetime import datetime, timedelta
from io import BytesIO

import pytest
import requests
from openpyxl import load_workbook

from conftest import BASE_URL


# ---------------- Auth module ----------------
class TestAuth:
    """Auth: login, signup pending, approval, reject"""

    def test_login_success(self, admin_client, test_credentials):
        r = admin_client.post(f"{BASE_URL}/api/auth/login", json=test_credentials, timeout=60)
        assert r.status_code == 200
        d = r.json()
        assert isinstance(d["token"], str) and len(d["token"]) > 20
        assert d["user"]["email"] == test_credentials["email"].lower()
        assert d["user"]["role"] == "admin"
        assert d["user"]["status"] == "approved"
        assert "password_hash" not in d["user"]

    def test_login_wrong_password(self, api_client, test_credentials):
        r = api_client.post(f"{BASE_URL}/api/auth/login",
                            json={"email": test_credentials["email"], "password": "wrong-xyz"}, timeout=60)
        assert r.status_code == 401
        assert "detail" in r.json()

    def test_me_requires_token(self, api_client):
        r = api_client.get(f"{BASE_URL}/api/auth/me", timeout=60)
        assert r.status_code == 401

    def test_me_rejects_bad_token(self, api_client):
        r = api_client.get(f"{BASE_URL}/api/auth/me",
                           headers={"Authorization": "Bearer not.a.token"}, timeout=60)
        assert r.status_code == 401

    def test_signup_pending_then_approve_then_login_and_reject(self, api_client, admin_client):
        email = f"test_qa_{uuid.uuid4().hex[:8]}@example.com"
        pw = "Test12345"
        r = api_client.post(f"{BASE_URL}/api/auth/signup",
                            json={"name": "TEST_QA User", "email": email, "password": pw}, timeout=60)
        assert r.status_code == 200
        assert r.json()["status"] == "pending"

        # duplicate signup rejected
        r_dup = api_client.post(f"{BASE_URL}/api/auth/signup",
                                json={"name": "TEST_QA User", "email": email, "password": pw}, timeout=60)
        assert r_dup.status_code == 400

        # pending cannot login
        r = api_client.post(f"{BASE_URL}/api/auth/login", json={"email": email, "password": pw}, timeout=60)
        assert r.status_code == 403
        assert "approval" in str(r.json()["detail"]).lower()

        # admin sees the user
        users = admin_client.get(f"{BASE_URL}/api/admin/users", timeout=60).json()
        u = next((x for x in users if x["email"] == email), None)
        assert u is not None and u["status"] == "pending"
        assert "password_hash" not in u

        # approve
        r = admin_client.post(f"{BASE_URL}/api/admin/users/{u['id']}/approve", timeout=60)
        assert r.status_code == 200 and r.json()["success"] is True

        # login now works
        r = api_client.post(f"{BASE_URL}/api/auth/login", json={"email": email, "password": pw}, timeout=60)
        assert r.status_code == 200
        op_token = r.json()["token"]

        # non-admin cannot access admin endpoints
        r = requests.get(f"{BASE_URL}/api/admin/users",
                         headers={"Authorization": f"Bearer {op_token}"}, timeout=60)
        assert r.status_code == 403

        # reject / delete account
        r = admin_client.delete(f"{BASE_URL}/api/admin/users/{u['id']}", timeout=60)
        assert r.status_code == 200
        users = admin_client.get(f"{BASE_URL}/api/admin/users", timeout=60).json()
        assert all(x["email"] != email for x in users)

    def test_signup_short_password_rejected(self, api_client):
        r = api_client.post(f"{BASE_URL}/api/auth/signup",
                            json={"name": "TEST_QA", "email": f"test_qa_{uuid.uuid4().hex[:6]}@ex.com",
                                  "password": "123"}, timeout=60)
        assert r.status_code == 422

    def test_admin_cannot_be_deleted(self, admin_client, test_credentials):
        users = admin_client.get(f"{BASE_URL}/api/admin/users", timeout=60).json()
        admin = next(x for x in users if x["email"] == test_credentials["email"].lower())
        r = admin_client.delete(f"{BASE_URL}/api/admin/users/{admin['id']}", timeout=60)
        assert r.status_code == 200
        users2 = admin_client.get(f"{BASE_URL}/api/admin/users", timeout=60).json()
        assert any(x["email"] == test_credentials["email"].lower() for x in users2), "admin was deleted!"

    def test_brute_force_lockout(self, api_client):
        """Playbook: lock after 5 failed attempts. Uses a dummy email so the real admin is never locked."""
        email = "TEST_QA_brute_bt@example.com"
        codes = []
        for _ in range(6):
            r = api_client.post(f"{BASE_URL}/api/auth/login",
                                json={"email": email, "password": "bad-pass"}, timeout=60)
            codes.append(r.status_code)
        assert codes[:5] == [401] * 5, f"unexpected codes before lockout: {codes}"
        assert codes[5] == 423, f"no lockout on 6th failed login, codes={codes}"
        body = api_client.post(f"{BASE_URL}/api/auth/login",
                               json={"email": email, "password": "bad-pass"}, timeout=60).json()
        msg = body.get("detail") or ""
        assert "dikunci" in msg.lower() or "terlalu banyak" in msg.lower(), f"non-Indonesian lock message: {msg}"


# ---------------- Owners master data ----------------
class TestOwners:
    def test_seed_owners_present(self, owners_list):
        names = {o["name"]: o for o in owners_list}
        for n, cap, veh in [("Ita Sari", 1300, "Pick Up"), ("Epit", 600, "Tossa"),
                            ("Sita Rosiani", 1800, "Hilux"), ("Suparmin", 1800, "Hilux")]:
            assert n in names, f"seed owner {n} missing"
            assert names[n]["capacity_kg"] == cap
            assert names[n]["vehicle_type"] == veh
        assert all("_id" not in o for o in owners_list)

    def test_create_and_update_owner(self, admin_client):
        payload = {"name": "TEST_QA Owner", "vehicle_type": "Truk", "capacity_kg": 2500, "active": True}
        r = admin_client.post(f"{BASE_URL}/api/owners", json=payload, timeout=60)
        assert r.status_code == 200
        oid = r.json()["id"]
        assert isinstance(oid, str)

        got = admin_client.get(f"{BASE_URL}/api/owners", timeout=60).json()
        created = next(o for o in got if o["id"] == oid)
        assert created["name"] == payload["name"] and created["capacity_kg"] == 2500

        upd = {**payload, "id": oid, "capacity_kg": 2700, "name": "TEST_QA Owner Edited"}
        r = admin_client.post(f"{BASE_URL}/api/owners", json=upd, timeout=60)
        assert r.status_code == 200
        got = admin_client.get(f"{BASE_URL}/api/owners", timeout=60).json()
        edited = next(o for o in got if o["id"] == oid)
        assert edited["capacity_kg"] == 2700 and edited["name"] == "TEST_QA Owner Edited"
        assert len([o for o in got if o["id"] == oid]) == 1

    def test_owners_requires_auth(self, api_client):
        assert api_client.get(f"{BASE_URL}/api/owners", timeout=60).status_code == 401


# ---------------- Transaction CRUD ----------------
class TestTransactions:
    DATE_A = "2026-11-03"
    DATE_B = "2026-11-04"

    def _payload(self, batch_id, owner, date, **kw):
        d = {"batch_id": batch_id, "date": date, "arrival_time": "08:10:05", "exit_time": "08:38:20",
             "owner_id": owner["id"], "owner_name": owner["name"], "vehicle_type": owner["vehicle_type"],
             "gross_kg": 1200, "tare_kg": None, "price_per_kg": 3100, "note": "TEST_QA"}
        d.update(kw)
        return d

    @pytest.fixture(scope="class", autouse=True)
    def cleanup(self, admin_client, active_batch):
        yield
        for d in (self.DATE_A, self.DATE_B):
            admin_client.delete(f"{BASE_URL}/api/transactions/date/{d}?batch_id={active_batch}", timeout=60)

    def test_create_computes_netto_and_persists(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Sita Rosiani")
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1500, tare_kg=200),
                              timeout=60)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["netto_kg"] == 1300
        assert d["netto_ton"] == 1.30
        assert d["total_amount"] == 1300 * 3100
        assert d["transaction_code"].startswith("SJM-202611-")
        tid, code = d["id"], d["transaction_code"]

        rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        got = next(x for x in rows if x["id"] == tid)
        assert got["netto_kg"] == 1300 and got["total_amount"] == 1300 * 3100
        assert "_id" not in got

        # edit must keep transaction_code
        upd = self._payload(active_batch, owner, self.DATE_A, gross_kg=1700, tare_kg=100)
        upd.update({"id": tid, "transaction_code": code})
        r = admin_client.post(f"{BASE_URL}/api/transactions", json=upd, timeout=60)
        assert r.status_code == 200
        e = r.json()
        assert e["transaction_code"] == code, "transaction_code changed on edit"
        assert e["netto_kg"] == 1600 and e["total_amount"] == 1600 * 3100

        rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        got = next(x for x in rows if x["id"] == tid)
        assert got["netto_kg"] == 1600 and got["transaction_code"] == code

        # delete single
        assert admin_client.delete(f"{BASE_URL}/api/transactions/{tid}", timeout=60).status_code == 200
        rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        assert all(x["id"] != tid for x in rows)

    def test_empty_tare_treated_as_zero(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Epit")
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=550, tare_kg=None,
                                                 arrival_time="09:05:00", exit_time="09:30:00"), timeout=60)
        assert r.status_code == 200
        d = r.json()
        assert d["netto_kg"] == 550 and d["netto_ton"] == 0.55

    def test_gross_over_capacity_rejected(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Epit")  # 600 kg
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1200), timeout=60)
        assert r.status_code == 400
        assert "kapasitas" in str(r.json()["detail"]).lower()

    def test_lunch_break_rejected(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Ita Sari")
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1000,
                                                 arrival_time="12:30:00", exit_time="12:55:00"), timeout=60)
        assert r.status_code == 400
        assert "istirahat" in str(r.json()["detail"]).lower()

    def test_outside_hours_rejected(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Ita Sari")
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1000,
                                                 arrival_time="17:30:00", exit_time="17:55:00"), timeout=60)
        assert r.status_code == 400

    def test_exit_before_arrival_rejected(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Ita Sari")
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1000,
                                                 arrival_time="10:30:00", exit_time="10:00:00"), timeout=60)
        assert r.status_code == 400

    def test_unknown_owner_rejected(self, admin_client, active_batch, owners_list):
        owner = {"id": str(uuid.uuid4()), "name": "Ghost", "vehicle_type": "X"}
        r = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A), timeout=60)
        assert r.status_code == 400

    def test_delete_by_date_only_that_date(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Suparmin")
        a = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_A, gross_kg=1700,
                                                 arrival_time="10:05:00", exit_time="10:30:00"), timeout=60)
        b = admin_client.post(f"{BASE_URL}/api/transactions",
                              json=self._payload(active_batch, owner, self.DATE_B, gross_kg=1600,
                                                 arrival_time="10:05:00", exit_time="10:30:00"), timeout=60)
        assert a.status_code == 200 and b.status_code == 200
        id_b = b.json()["id"]

        r = admin_client.delete(f"{BASE_URL}/api/transactions/date/{self.DATE_A}?batch_id={active_batch}", timeout=60)
        assert r.status_code == 200
        assert r.json()["deleted_count"] >= 1 and r.json()["date"] == self.DATE_A

        rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        assert all(x["date"] != self.DATE_A for x in rows), "date A rows remain"
        assert any(x["id"] == id_b for x in rows), "date B rows wrongly deleted"

        admin_client.delete(f"{BASE_URL}/api/transactions/date/{self.DATE_B}?batch_id={active_batch}", timeout=60)


# ---------------- Generator (atomic, exact target, schedule rules) ----------------
class TestGenerator:
    START = "2026-07-31"
    END = "2026-08-01"

    @pytest.fixture(scope="class", autouse=True)
    def cleanup(self, admin_client, active_batch):
        yield
        for d in (self.START, self.END):
            admin_client.delete(f"{BASE_URL}/api/transactions/date/{d}?batch_id={active_batch}", timeout=60)

    def test_infeasible_returns_422_and_is_atomic(self, admin_client, active_batch):
        # count only rows for the generator window to stay isolated from parallel test classes
        def window_rows():
            rows = admin_client.get(f"{BASE_URL}/api/transactions", timeout=60).json()
            return [x for x in rows if x["date"] == "2026-12-01"]

        before = window_rows()
        r = admin_client.post(f"{BASE_URL}/api/generator/run",
                              json={"start_date": "2026-12-01", "end_date": "2026-12-01",
                                    "target_kg": 900000, "price_per_kg": 3000}, timeout=180)
        assert r.status_code == 422, r.text
        d = r.json()["detail"]
        for k in ("target", "maximum", "shortage", "reason", "status", "message"):
            assert k in d, f"missing key {k}"
        assert d["status"] == "infeasible"
        assert d["target"] == 900000
        assert d["shortage"] == d["target"] - d["maximum"]
        after = window_rows()
        assert len(after) == len(before) == 0, "transactions were inserted despite infeasible target"

    def test_invalid_inputs(self, admin_client):
        r = admin_client.post(f"{BASE_URL}/api/generator/run",
                              json={"start_date": "2026-08-05", "end_date": "2026-08-01",
                                    "target_kg": 1000, "price_per_kg": 3000}, timeout=60)
        assert r.status_code == 400
        r = admin_client.post(f"{BASE_URL}/api/generator/run",
                              json={"start_date": "2026-08-01", "end_date": "2026-08-01",
                                    "target_kg": 0, "price_per_kg": 3000}, timeout=60)
        assert r.status_code == 400

    def test_feasible_exact_target_and_rules(self, admin_client, active_batch, owners_list):
        target = 6450
        r = admin_client.post(f"{BASE_URL}/api/generator/run",
                              json={"start_date": self.START, "end_date": self.END,
                                    "target_kg": target, "price_per_kg": 3100}, timeout=180)
        assert r.status_code == 200, r.text
        res = r.json()
        assert res["actual_kg"] == target, f"actual {res['actual_kg']} != target {target}"
        assert res["status"] == "exact"
        assert res["generated_count"] > 0

        rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        gen = [x for x in rows if x["date"] in (self.START, self.END)]
        assert len(gen) == res["generated_count"]
        assert sum(x["netto_kg"] for x in gen) == target
        caps = {o["id"]: o["capacity_kg"] for o in owners_list}

        codes = [x["transaction_code"] for x in gen]
        assert len(set(codes)) == len(codes), f"duplicate transaction codes: {codes}"

        for x in gen:
            assert x["gross_kg"] <= caps[x["owner_id"]], f"gross over capacity: {x}"
            assert x["total_amount"] == x["netto_kg"] * 3100
            assert x["netto_ton"] == round(x["netto_kg"] / 1000, 2)
            arr = datetime.strptime(x["arrival_time"], "%H:%M:%S")
            ex = datetime.strptime(x["exit_time"], "%H:%M:%S")
            assert datetime.strptime("07:00:00", "%H:%M:%S") <= arr <= datetime.strptime("16:00:00", "%H:%M:%S"), x
            if datetime.strptime("12:00:00", "%H:%M:%S") <= arr < datetime.strptime("13:00:00", "%H:%M:%S"):
                assert "istirahat" in x["note"].lower(), f"lunch slot without note: {x}"
            delta = (ex - arr).total_seconds() / 60
            assert 20 <= delta <= 41, f"weigh duration {delta} out of range: {x}"

        for day in (self.START, self.END):
            same = sorted([x for x in gen if x["date"] == day], key=lambda t: t["arrival_time"])
            prev = None
            last_owner_arr = {}
            for x in same:
                arr = datetime.strptime(f"{day} {x['arrival_time']}", "%Y-%m-%d %H:%M:%S")
                if prev:
                    prev_exit = datetime.strptime(f"{day} {prev['exit_time']}", "%Y-%m-%d %H:%M:%S")
                    assert arr > prev_exit, f"overlap: {prev} -> {x}"
                    assert prev["owner_id"] != x["owner_id"], f"consecutive same owner: {prev} -> {x}"
                if x["owner_id"] in last_owner_arr:
                    gap = (arr - last_owner_arr[x["owner_id"]]).total_seconds() / 60
                    assert gap >= 105, f"owner gap {gap} min < 105: {x}"
                last_owner_arr[x["owner_id"]] = arr
                prev = x


# ---------------- Analytics + Excel export ----------------
class TestAnalyticsExport:
    DATE = "2026-11-10"

    @pytest.fixture(scope="class", autouse=True)
    def seeded(self, admin_client, active_batch, owners_list):
        owner = next(o for o in owners_list if o["name"] == "Sita Rosiani")
        payloads = [
            {"gross_kg": 1700, "tare_kg": 200, "arrival_time": "08:00:00", "exit_time": "08:25:00"},
            {"gross_kg": 1500, "tare_kg": None, "arrival_time": "11:00:00", "exit_time": "11:25:00"},
        ]
        for p in payloads:
            r = admin_client.post(f"{BASE_URL}/api/transactions", json={
                "batch_id": active_batch, "date": self.DATE, "owner_id": owner["id"],
                "owner_name": owner["name"], "vehicle_type": owner["vehicle_type"],
                "price_per_kg": 3000, "note": "TEST_QA", **p}, timeout=60)
            assert r.status_code == 200, r.text
        yield
        admin_client.delete(f"{BASE_URL}/api/transactions/date/{self.DATE}?batch_id={active_batch}", timeout=60)

    def test_summary_totals_and_filters(self, admin_client, active_batch):
        r = admin_client.get(
            f"{BASE_URL}/api/analytics/summary?batch_id={active_batch}&start={self.DATE}&end={self.DATE}", timeout=60)
        assert r.status_code == 200
        s = r.json()
        assert s["total_transactions"] == 2
        assert s["total_netto_kg"] == 1500 + 1500
        assert s["total_netto_ton"] == 3.0
        assert s["total_spending"] == 3000 * 3000
        assert len(s["daily_summary"]) == 1 and s["daily_summary"][0]["label"] == self.DATE
        assert any(o["label"] == "Sita Rosiani" for o in s["owner_summary"])

        r2 = admin_client.get(
            f"{BASE_URL}/api/analytics/summary?batch_id={active_batch}&start=2026-01-01&end=2026-01-02", timeout=60)
        assert r2.json()["total_transactions"] == 0

    def test_excel_export_contains_formulas(self, admin_client, active_batch):
        r = admin_client.get(
            f"{BASE_URL}/api/export/excel?batch_id={active_batch}&start={self.DATE}&end={self.DATE}", timeout=120)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")
        wb = load_workbook(BytesIO(r.content))
        # Workbook multi-sheet: buku penerimaan + rekap pemilik (per orang) + harian + laba rugi + harga
        assert "Rekap Pemilik" in wb.sheetnames, wb.sheetnames
        assert "Rekap Harian" in wb.sheetnames, wb.sheetnames
        assert "Laba Rugi" in wb.sheetnames, wb.sheetnames
        assert "Harga Harian" in wb.sheetnames, wb.sheetnames
        ws = wb["Buku Penerimaan"]
        assert "BUKU PENERIMAAN" in str(ws["A1"].value)
        hr = next(row for row in range(1, 12) if ws.cell(row=row, column=1).value == "Kode")
        first = hr + 1
        assert ws.cell(row=first, column=10).value == f"=H{first}-I{first}"
        assert ws.cell(row=first, column=11).value == f"=ROUND(J{first}/1000,2)"
        assert ws.cell(row=first, column=13).value == f"=J{first}*L{first}"
        last = first + 1  # dua transaksi pada tanggal test
        total_row = last + 1
        assert ws.cell(row=total_row, column=1).value == "TOTAL"
        assert ws.cell(row=total_row, column=8).value == f"=SUM(H{first}:H{last})"
        assert ws.cell(row=total_row, column=10).value == f"=SUM(J{first}:J{last})"
        assert ws.cell(row=total_row, column=13).value == f"=SUM(M{first}:M{last})"

    def test_owner_daily_finance_exports(self, admin_client, active_batch):
        for path, sheet in (("owners-excel", "Ringkasan Pemilik"),
                            ("daily-excel", "Rekap Harian"),
                            ("finance-excel", "Laba Rugi Harian")):
            r = admin_client.get(f"{BASE_URL}/api/export/{path}?batch_id={active_batch}", timeout=120)
            assert r.status_code == 200, (path, r.status_code)
            assert "spreadsheetml" in r.headers.get("content-type", "")
            wb = load_workbook(BytesIO(r.content))
            assert sheet in wb.sheetnames, (path, wb.sheetnames)

    def test_owner_export_single_person(self, admin_client, active_batch):
        txs = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={active_batch}", timeout=60).json()
        owner = txs[0]["owner_name"]
        r = admin_client.get(f"{BASE_URL}/api/export/owners-excel?batch_id={active_batch}&owner={owner}", timeout=120)
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert wb.sheetnames == [owner[:31]], wb.sheetnames
        bad = admin_client.get(f"{BASE_URL}/api/export/owners-excel?batch_id={active_batch}&owner=Tidak%20Ada", timeout=60)
        assert bad.status_code == 404

    def test_prices_crud_and_export(self, admin_client):
        for date, price, note in (("2026-04-01", 2500, "awal"), ("2026-04-02", 2700, "naik"), ("2026-04-03", 2600, "turun")):
            assert admin_client.post(f"{BASE_URL}/api/prices", json={
                "date": date, "price_per_kg": price, "note": note}, timeout=30).status_code == 200
        data = admin_client.get(f"{BASE_URL}/api/prices?month=2026-04", timeout=30).json()
        rows = {p["date"]: p for p in data["prices"]}
        assert rows["2026-04-01"]["change"] is None and rows["2026-04-01"]["trend"] == "awal"
        assert rows["2026-04-02"]["change"] == 200 and rows["2026-04-02"]["trend"] == "naik"
        assert rows["2026-04-03"]["change"] == -100 and rows["2026-04-03"]["trend"] == "turun"
        assert data["stats"]["high"] == 2700 and data["stats"]["low"] == 2500
        r = admin_client.get(f"{BASE_URL}/api/export/prices-excel?month=2026-04", timeout=60)
        assert r.status_code == 200 and "Harga Harian" in load_workbook(BytesIO(r.content)).sheetnames
        assert admin_client.post(f"{BASE_URL}/api/prices", json={"date": "2026-04-04", "price_per_kg": 0}, timeout=30).status_code == 400
        for date in ("2026-04-01", "2026-04-02", "2026-04-03"):
            assert admin_client.delete(f"{BASE_URL}/api/prices/{date}", timeout=30).status_code == 200
        assert admin_client.delete(f"{BASE_URL}/api/prices/2026-04-01", timeout=30).status_code == 404

    def test_export_requires_auth(self, api_client):
        assert api_client.get(f"{BASE_URL}/api/export/excel", timeout=60).status_code == 401


# ---------------- Batches ----------------
class TestBatches:
    def test_new_batch_archives_previous_and_keeps_old_data(self, admin_client, owners_list):
        b0 = admin_client.get(f"{BASE_URL}/api/batches", timeout=60).json()
        old_active = next(b for b in b0 if b["active"])
        owner = next(o for o in owners_list if o["name"] == "Ita Sari")
        date = "2026-11-20"
        r = admin_client.post(f"{BASE_URL}/api/transactions", json={
            "batch_id": old_active["id"], "date": date, "arrival_time": "09:00:00", "exit_time": "09:25:00",
            "owner_id": owner["id"], "owner_name": owner["name"], "vehicle_type": owner["vehicle_type"],
            "gross_kg": 1200, "tare_kg": None, "price_per_kg": 3000, "note": "TEST_QA"}, timeout=60)
        assert r.status_code == 200
        old_tx_id = r.json()["id"]

        rnew = admin_client.post(f"{BASE_URL}/api/batches/new", timeout=60)
        assert rnew.status_code == 200
        new_id = rnew.json()["id"]
        assert new_id != old_active["id"]

        b1 = admin_client.get(f"{BASE_URL}/api/batches", timeout=60).json()
        assert [b for b in b1 if b["active"]][0]["id"] == new_id
        assert next(b for b in b1 if b["id"] == old_active["id"])["active"] is False
        assert all("_id" not in b for b in b1)

        # new batch empty, old batch data intact
        assert admin_client.get(f"{BASE_URL}/api/transactions?batch_id={new_id}", timeout=60).json() == []
        old_rows = admin_client.get(f"{BASE_URL}/api/transactions?batch_id={old_active['id']}", timeout=60).json()
        assert any(x["id"] == old_tx_id for x in old_rows)

        # cleanup: remove test tx from archived batch
        admin_client.delete(f"{BASE_URL}/api/transactions/{old_tx_id}", timeout=60)
