"""Verifikasi PPh 22 yang bisa dikelola: default tarif, tarif per tanggal, nominal manual, reset."""
import io, requests
from openpyxl import load_workbook

BASE = 'http://localhost:8001'
s = requests.Session()
s.headers['Authorization'] = 'Bearer ' + s.post(f'{BASE}/api/auth/login', json={
    'email': 'yumaclovstar@gmail.com', 'password': '178910'}).json()['token']

b = next(x for x in s.get(f'{BASE}/api/batches').json() if x['active'])['id']
if not s.get(f'{BASE}/api/transactions?batch_id={b}').json():
    print('generator', s.post(f'{BASE}/api/generator/run', json={
        'start_date': '2026-09-01', 'end_date': '2026-09-02', 'target_kg': 8000, 'price_per_kg': 2900}).status_code)
txs = s.get(f'{BASE}/api/transactions?batch_id={b}').json()
date = sorted({t['date'] for t in txs})[0]
print('grading', s.post(f'{BASE}/api/finance/day', json={
    'batch_id': b, 'date': date, 'total_tare_kg': 200, 'sip_price_per_kg': 3500, 'freight_per_kg': 200}).status_code)


def row(d=date):
    return next(x for x in s.get(f'{BASE}/api/finance/summary?batch_id={b}').json()['days'] if x['date'] == d)


def totals():
    return s.get(f'{BASE}/api/finance/summary?batch_id={b}').json()['totals']


r = row()
hj, modal = r['harga_jual'], r['total_modal']
print('default rate', s.get(f'{BASE}/api/settings/pph').json())
assert r['pph22'] == round(hj * 0.0025) and r['pph_mode'] == 'auto' and r['pph_rate_pct'] == 0.25, r
assert r['untung_rugi'] == hj - r['pph22'] - modal

# 1) ubah tarif default global
print('put default 0.5%', s.put(f'{BASE}/api/settings/pph', json={'rate_pct': 0.5}).status_code)
r = row(); assert r['pph22'] == round(hj * 0.005) and r['pph_rate_pct'] == 0.5, r
assert r['untung_rugi'] == hj - r['pph22'] - modal

# 2) tarif khusus per tanggal (auto)
print('day rate 1.5%', s.post(f'{BASE}/api/finance/pph', json={'batch_id': b, 'date': date, 'mode': 'auto', 'rate_pct': 1.5, 'note': 'tarif khusus'}).json())
r = row(); assert r['pph22'] == round(hj * 0.015) and r['pph_custom'] and r['pph_note'] == 'tarif khusus', r

# 3) nominal manual dari hasil akhir
print('manual 123456', s.post(f'{BASE}/api/finance/pph', json={'batch_id': b, 'date': date, 'mode': 'manual', 'amount': 123456, 'note': 'bukti potong SIP'}).json())
r = row(); assert r['pph22'] == 123456 and r['pph_mode'] == 'manual', r
assert r['untung_rugi'] == hj - 123456 - modal
t = totals(); print('totals pph', t['pph22'], 'effective', t['pph_effective_pct'], 'manual days', t['pph_manual_days'])
assert t['pph22'] >= 123456 and t['pph_manual_days'] == 1

# validasi: manual > harga jual ditolak, tarif > 100 ditolak, tanggal belum grading ditolak
print('reject over', s.post(f'{BASE}/api/finance/pph', json={'batch_id': b, 'date': date, 'mode': 'manual', 'amount': hj + 1}).status_code,
      s.post(f'{BASE}/api/finance/pph', json={'batch_id': b, 'date': date, 'mode': 'auto', 'rate_pct': 120}).status_code,
      s.post(f'{BASE}/api/finance/pph', json={'batch_id': b, 'date': '2030-01-01', 'mode': 'auto'}).status_code,
      s.put(f'{BASE}/api/settings/pph', json={'rate_pct': -5}).status_code)

# excel memuat kolom tarif/mode
x = s.get(f'{BASE}/api/export/finance-excel?batch_id={b}')
wb = load_workbook(io.BytesIO(x.content))['Laba Rugi Harian']
hdr = [c.value for c in wb[next(r0 for r0 in range(1, 12) if wb.cell(row=r0, column=1).value == 'Tanggal')]]
print('excel headers', hdr)
assert 'Tarif / Mode PPh' in hdr and 'PPh 22' in hdr

# 4) reset ke default
print('reset', s.delete(f'{BASE}/api/finance/pph/{date}?batch_id={b}').json())
r = row(); assert r['pph22'] == round(hj * 0.005) and r['pph_mode'] == 'auto' and not r['pph_custom'], r
s.put(f'{BASE}/api/settings/pph', json={'rate_pct': 0.25})
r = row(); assert r['pph22'] == round(hj * 0.0025), r
print('ALL PPH CHECKS PASSED')
